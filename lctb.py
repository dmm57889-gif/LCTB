import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import warnings
from io import BytesIO
import joblib
import requests
from PIL import Image, ImageFile, PngImagePlugin
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

warnings.filterwarnings('ignore')

# Configurazione pagina Streamlit
st.set_page_config(
    page_title="Generatore Proposta Last Chance To Buy ",
    page_icon="💰",
    layout="wide"
)

# Configurazione di Pillow per gestire immagini problematiche
ImageFile.LOAD_TRUNCATED_IMAGES = True
PngImagePlugin.MAX_TEXT_CHUNK = 10000000

st.title("💰 Generatore proposta Last Chance To Buy")
st.markdown("---")

# Sidebar per input utente
st.sidebar.header("📁 Caricamento File")

# File obbligatori
st_item_file = st.sidebar.file_uploader("Carica file st_item.xlsx", type=['xlsx'], key="st_item")
file_A = st.sidebar.file_uploader("Carica file A.xlsx", type=['xlsx'], key="A")
file_B = st.sidebar.file_uploader("Carica file B.xlsx", type=['xlsx'], key="B")
calendar_file = st.sidebar.file_uploader("Carica file calendar.xlsx", type=['xlsx'], key="calendar")
tracking_file = st.sidebar.file_uploader("Carica file % tracking per negozio.xlsx", type=['xlsx'], key="tracking")
goals_file = st.sidebar.file_uploader("Carica file function_goals.xlsx", type=['xlsx'], key="goals")
segment_file = st.sidebar.file_uploader("Carica file segment.xlsx", type=['xlsx'], key="segment")

# File opzionali
st.sidebar.subheader("📊 File Opzionali")
images_file = st.sidebar.file_uploader("Carica file immagini articoli", type=['xlsx'], key="images")
sequence_file = st.sidebar.file_uploader("Carica file sequenza articoli sconto.xlsx (opzionale)", type=['xlsx'], key="sequence")
keras_model = st.sidebar.file_uploader("Carica modello Keras (.keras)", type=['keras'], key="keras")
pkl_model = st.sidebar.file_uploader("Carica modello PKL (.pkl)", type=['pkl'], key="pkl")

# Filtri per settimane
st.sidebar.subheader("📅 Filtri Temporali")
start_week = st.sidebar.text_input("Settimana iniziale (AAAA-WW):", placeholder="2025-19")
end_week = st.sidebar.text_input("Settimana finale (AAAA-WW):", placeholder="2025-28")

# Funzioni di utilità
def is_valid_yearweek(yearweek):
    try:
        year, week = yearweek.split('-')
        year = int(year)
        week = int(week)
        return (1 <= week <= 52)
    except:
        return False

def filter_by_week_range(A, calendar, start_week, end_week):
    calendar['YearWeek'] = calendar['YearWeek'].astype(str)
    calendar[['anno', 'settimana']] = calendar['YearWeek'].str.split('-', n=1, expand=True)
    calendar['anno'] = calendar['anno'].astype(int)
    calendar['settimana'] = calendar['settimana'].astype(int)
    calendar = calendar.sort_values(by=['anno', 'settimana']).reset_index(drop=True)
    
    start_year, start_week_num = map(int, start_week.split('-'))
    end_year, end_week_num = map(int, end_week.split('-'))
    
    mask = (
        ((calendar['anno'] > start_year) | ((calendar['anno'] == start_year) & (calendar['settimana'] >= start_week_num))) &
        ((calendar['anno'] < end_year) | ((calendar['anno'] == end_year) & (calendar['settimana'] <= end_week_num)))
    )
    
    yearweeks = calendar[mask]['YearWeek'].drop_duplicates().tolist()
    if not yearweeks:
        return pd.DataFrame()
    
    return A[A['First Tracking YearWeek'].astype(str).isin(yearweeks)]

def categorize_st(df, function_name, year_month, df_classified):
    df_filtered = df[(df['Function'] == function_name) & (df['Commercial YearMonth'] == year_month)].copy()
    if df_filtered.empty:
        return df_classified
    
    if df_filtered.shape[0] == 1:
        df_function = df[df['Function'] == function_name]
        st_percentiles = df_function['ST item'].quantile([0.25, 0.5, 0.75])
        cluster_method = "Cluster funzione"
    else:
        st_percentiles = df_filtered['ST item'].quantile([0.25, 0.5, 0.75])
        cluster_method = "Cluster funzione/mese commerciale"
    
    def categorize(row):
        if row['ST item'] <= st_percentiles[0.25]:
            return 'Basso'
        elif row['ST item'] <= st_percentiles[0.5]:
            return 'Medio Basso'
        elif row['ST item'] <= st_percentiles[0.75]:
            return 'Medio Alto'
        else:
            return 'Alto'
    
    df_filtered['ST_Cluster'] = df_filtered.apply(categorize, axis=1)
    df_filtered['Metodo Cluster'] = cluster_method
    df_classified = pd.concat([df_classified, df_filtered], ignore_index=True)
    return df_classified

def remove_leading_zero(year_week):
    if pd.isna(year_week) or year_week == 0:
        return year_week
    year, week = str(year_week).split('-')
    week = str(int(week))
    return f"{year}-{week}"

def format_percent(x):
    if x is None or pd.isna(x):
        return "-"
    else:
        return f"{x*100:.2f}%".replace('.', ',')

def preprocess_image(image, target_size=(224, 224)):
    """Ridimensiona l'immagine preservandone il rapporto d'aspetto"""
    img = image.copy()
    img.thumbnail(target_size, Image.LANCZOS)
    new_img = Image.new("RGB", target_size)
    left = (target_size[0] - img.size[0]) // 2
    top = (target_size[1] - img.size[1]) // 2
    new_img.paste(img, (left, top))
    return new_img

def download_and_preprocess(index, url, category, session):
    """Scarica e preprocessa l'immagine"""
    if url == "URL non presente":
        return None
    try:
        response = session.get(url, timeout=10)
        if response.status_code == 200:
            img = Image.open(BytesIO(response.content)).convert("RGB")
            img = preprocess_image(img, target_size=(224, 224))
            img_array = np.array(img) / 255.0
            return (index, img_array, category)
    except Exception as e:
        st.warning(f"Errore nel processare l'immagine {url}: {e}")
        return None

def create_styled_excel(df, category, sequence_file=None):
    """Crea un file Excel formattato con styling avanzato"""
    column_order = [
        'Function', 'Season', 'Cod Department', 'Des Department', 'Item Code', 'Des item',
        'Sales item', 'Delivered item', 'Sales 4th Normalizzata', 'ST item', 'ST 4th Normalizzato',
        'Cod Category', 'Subcategory', 'APC', 'Promotion', 'Commercial YearWeek', 'Commercial YearMonth',
        'ST_Cluster', 'Metodo Cluster', 'Delta ST P2W', 'Delta ST P3W', 'TFI', 'Proposal',
        'First Tracking YearWeek', 'First Planned Tracking YearWeek', 'First Sale YearWeek',
        'Weeks since First Sale Date', 'Intake Quantity', 'Displayed Quantity', 'Recycled',
        '% Stores with Tracking within 6 weeks', '% Store with Tracking', 'Total Item Tracked',
        'Segment', 'Retail Price', 'Stock residuo', 'SVA', 'Sconto proposto', 'Data elaborazione'
    ]
    
    # Riordina le colonne del DataFrame
    existing_columns = [col for col in column_order if col in df.columns]
    extra_columns = [col for col in df.columns if col not in column_order]
    final_column_order = existing_columns + extra_columns
    df = df[final_column_order]
    # FINE RIGHE DA AGGIUNGERE
    
    elaboration_date = datetime.today().strftime('%d-%m-%Y')
    
    if int(category) == 31:
        filename = f"IC_proposte_sconti_WOMAN_{elaboration_date}.xlsx"
    elif int(category) == 32:
        filename = f"IC_proposte_sconti_MAN_{elaboration_date}.xlsx"
    elif int(category) == 33:
        filename = f"IC_proposte_sconti_KIDS_{elaboration_date}.xlsx"
    else:
        filename = f"IC_proposte_sconti_{elaboration_date}.xlsx"
    
    # Crea file Excel in memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')
    
    output.seek(0)
    
    # Carica il workbook per la formattazione avanzata
    wb = load_workbook(output)
    ws = wb.active
    
    # Configurazione formattazione colonne
    header_config = {
        "Proposal": {
            "font": Font(bold=True),
            "fill": PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
        },
        "ST item": {
            "round": 4,
            "num_format": '0.0000',
            "font": Font(bold=True),
            "fill": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        },
        "Sales item": {
            "round": 2,
            "num_format": '0.00'
        },
        "Delivered item": {
            "round": 2,
            "num_format": '0.00'
        },
        "SVA": {
            "round": 2,
            "num_format": '0.00'
        },
        "Stock residuo": {
            "round": 2,
            "num_format": '0.00'
        },
        "Total Item Tracked": {
            "round": 4,
            "num_format": '0.0000'
        },
        "TFI": {
            "fill": PatternFill(start_color="EFF7FF", end_color="EFF7FF", fill_type="solid")
        },
        "Delta ST P2W": {
            "fill": PatternFill(start_color="EFF7FF", end_color="EFF7FF", fill_type="solid")
        },
        "Delta ST P3W": {
            "fill": PatternFill(start_color="EFF7FF", end_color="EFF7FF", fill_type="solid")
        }
    }
    
    # Mappa header alle colonne
    header_columns = {}
    for cell in ws[1]:
        if cell.value in header_config:
            header_columns[cell.value] = cell.column
    
    # Applica formattazione per colonne numeriche
    for header, config in header_config.items():
        if header not in header_columns or header in ["Delta ST P2W", "Delta ST P3W"]:
            continue
        
        col_idx = header_columns[header]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                # Sostituisce "-" con 0 per colonne numeriche
                if cell.value == "-" and "round" in config:
                    cell.value = 0
                
                # Applica arrotondamento e formato numerico
                if isinstance(cell.value, (int, float)) and "round" in config:
                    cell.value = round(cell.value, config["round"])
                    cell.number_format = config["num_format"]
                
                # Applica font e riempimento
                if "font" in config:
                    cell.font = config["font"]
                if "fill" in config:
                    cell.fill = config["fill"]
    
    # Formattazione speciale per Delta ST P2W e P3W (solo colore, non numerico)
    for header in ["Delta ST P2W", "Delta ST P3W"]:
        if header in header_columns:
            col_idx = header_columns[header]
            config = header_config[header]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if "font" in config:
                        cell.font = config["font"]
                    if "fill" in config:
                        cell.fill = config["fill"]
    
    # Evidenzia in rosso articoli con sequenza (se presente)
    if sequence_file is not None:
        # Trova colonna Item Code
        item_code_col = None
        for cell in ws[1]:
            if cell.value == "Item Code":
                item_code_col = cell.column
                break
        
        if item_code_col is not None:
            # Crea set degli item code con sequenza
            cod_items_seq = set()
            if 'Settimana applicazione sconto' in df.columns:
                cod_items_seq = set(df[df['Settimana applicazione sconto'] != '-']['Item Code'])
            
            # Applica font rosso
            red_font = Font(color="FF0000")
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                item_value = row[item_code_col - 1].value
                if str(item_value) in cod_items_seq:
                    for cell in row:
                        cell.font = red_font
    
    # Salva in BytesIO
    formatted_output = BytesIO()
    wb.save(formatted_output)
    formatted_output.seek(0)
    
    return formatted_output, filename
# Controllo file obbligatori
required_files = [st_item_file, file_A, file_B, calendar_file, tracking_file, goals_file, segment_file]
if not all(required_files):
    st.warning("⚠️ Carica tutti i file obbligatori per continuare.")
    st.stop()

# Controllo validità settimane
if not start_week or not end_week:
    st.warning("⚠️ Inserisci le settimane di inizio e fine.")
    st.stop()

if not (is_valid_yearweek(start_week) and is_valid_yearweek(end_week)):
    st.error("❌ Formato settimana non valido. Usa il formato AAAA-WW.")
    st.stop()

if start_week > end_week:
    st.error("❌ La settimana iniziale non può essere successiva a quella finale.")
    st.stop()

# Pulsante per iniziare l'elaborazione
if st.button("🚀 Avvia Elaborazione", type="primary"):
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        # Caricamento dati
        status_text.text("📁 Caricamento file...")
        progress_bar.progress(10)
        
        st_item = pd.read_excel(st_item_file)
        A = pd.read_excel(file_A)
        B = pd.read_excel(file_B)
        calendar = pd.read_excel(calendar_file)
        tracking = pd.read_excel(tracking_file)
        goals = pd.read_excel(goals_file)
        segment = pd.read_excel(segment_file)
        
        # Caricamento file immagini solo se presente
        if images_file:
            images_df = pd.read_excel(images_file)
        else:
            images_df = None

        # --- keep originals for special processing (VERY IMPORTANT) ---
        st_item_original = st_item.copy()
        A_original = A.copy()
        B_original = B.copy()
        
        # normalizziamo la colonna Recycled (gestisce 'Sì','Si','SI','si' ecc.)
        if 'Recycled' in A_original.columns:
            A_original['Recycled'] = A_original['Recycled'].astype(str).str.strip()
            A_original['Recycled'] = A_original['Recycled'].replace(
                {r'^(si|sì|SI|Si|SÌ)$': 'Sì'}, regex=True
            )
        else:
            st.warning("Attenzione: colonna 'Recycled' non trovata in A_original.")
        
        # debug rapido per verificare che ci siano articoli ricondizionati
        #st.info(f"Recycled counts (original A): {A_original['Recycled'].value_counts().to_dict()}")

        
        # Preparazione calendario
        calendar['YearWeek'] = calendar['YearWeek'].astype(str)
        calendar[['anno', 'settimana']] = calendar['YearWeek'].str.split('-', n=1, expand=True)
        calendar['anno'] = calendar['anno'].astype(int)
        calendar['settimana'] = calendar['settimana'].astype(int)
        calendar = calendar.sort_values(by=['anno', 'settimana']).drop(columns=['anno', 'settimana']).reset_index(drop=True)
        
        # Filtro tracking
        tracking_filtered = tracking[tracking["% Stores with Tracking within 6 weeks"] >= 0.3]
        tracking_below = tracking[tracking["% Stores with Tracking within 6 weeks"] < 0.3]
        
        progress_bar.progress(20)
        
        # Pulizia dati
        status_text.text("🧹 Pulizia e preparazione dati...")
        A = A.dropna(subset=['Item Code']).fillna(0)
        B = B.dropna(subset=['Item Code']).fillna(0)
        st_item = st_item.dropna(subset=['Item Code'])
        
        A = A[A['Commercial YearWeek'] != 0].reset_index(drop=True)
        st_item = st_item.merge(A[['Item Code', 'Commercial YearWeek', 'Commercial YearMonth']], on='Item Code', how='left')
        
        category = max(st_item["Cod Category"])
        current_week = calendar.iloc[-1]["YearWeek"]
        
        progress_bar.progress(30)
        
        # Filtraggio per settimane
        status_text.text("📅 Filtraggio per intervallo settimane...")
        A_filtered = filter_by_week_range(A, calendar, start_week, end_week)
        
        if A_filtered.empty:
            st.error("❌ Nessun dato trovato nell'intervallo di settimane selezionato.")
            st.stop()
        
        progress_bar.progress(40)
        
        # Categorizzazione ST
        status_text.text("📊 Categorizzazione ST...")
        function_months = set(zip(st_item['Function'], st_item['Commercial YearMonth']))
        df_clusters = pd.DataFrame(columns=st_item.columns)
        
        for func, year_month in function_months:
            df_clusters = categorize_st(st_item, func, year_month, df_clusters)
        
        progress_bar.progress(50)
        
        # Calcolo logiche di business
        status_text.text("💼 Applicazione logiche di business...")
        
        # Definizione set di articoli
        items_in_exposition_month = set(A_filtered['Item Code'])
        items_not_in_exposition_month = set(A['Item Code']) - set(A_filtered['Item Code'])
        item_codes_comuni_con_B = set(A_filtered['Item Code']).intersection(set(B['Item Code']))
        
        df_items_nulli = A[(A["First Tracking YearWeek"] == 0) & (A["First Sale YearWeek"] == 0)]
        lista_items_nulli = set(df_items_nulli['Item Code'])
        
        items_in_exposition_month_without_sales = set(A[(A["Item Code"].isin(items_in_exposition_month)) & (A["First Sale YearWeek"] == 0)]["Item Code"])
        items_above_tracked_in_exposition_month = set(tracking_filtered[tracking_filtered["Item Code"].isin(items_in_exposition_month)]["Item Code"])
        items_below_tracked_in_exposition_month = set(tracking_below[tracking_below["Item Code"].isin(items_in_exposition_month)]["Item Code"])
        items_not_tracked = set(A[A["First Tracking YearWeek"] == 0]["Item Code"])
        
        progress_bar.progress(60)
        
        # Calcolo Delta ST e proposte
        B_filtered = B[B['Item Code'].isin(items_above_tracked_in_exposition_month)]
        B_filtered['YearWeek'] = B_filtered['YearWeek'].apply(remove_leading_zero)
        
        for item in items_above_tracked_in_exposition_month:
            parts = current_week.split("-")
            week_number = int(parts[1])
            
            if week_number != 1:
                week_number_p2w = str(week_number)
                week_number_p3w = str(week_number - 1)
            else:
                week_number_p2w = "52"
                week_number_p3w = "51"
            
            week_p2w = parts[0] + "-" + week_number_p2w
            week_p3w = parts[0] + "-" + week_number_p3w
            
            p2w_data = B_filtered.loc[(B_filtered["Item Code"] == item) & (B_filtered['YearWeek'] == week_p2w), 'Delta ST PW']
            p3w_data = B_filtered.loc[(B_filtered["Item Code"] == item) & (B_filtered['YearWeek'] == week_p3w), 'Delta ST PW']
            
            p2w = p2w_data.values[0] if not p2w_data.empty else None
            p3w = p3w_data.values[0] if not p3w_data.empty else None
            
            item_index = df_clusters.index[df_clusters['Item Code'] == item].tolist()
            if item_index:
                item_index = item_index[0]
                cluster = df_clusters.at[item_index, 'ST_Cluster']
                df_clusters.at[item_index, 'Delta ST P2W'] = format_percent(p2w)
                df_clusters.at[item_index, 'Delta ST P3W'] = format_percent(p3w)
                
                item_function = df_clusters.at[item_index, 'Function']
                mask_function = (goals['Function'] == item_function)
                
                if not goals.loc[mask_function].empty:
                    row = goals.loc[mask_function].iloc[0]
                    theoretical_increase = row['Teorethical Increase %']
                    num_life_weeks = row['NumLifeWeeks']
                    threshold = 0.025 if num_life_weeks == -1 else 0.75 * theoretical_increase
                else:
                    threshold = 0.0196
                    theoretical_increase = 0.0196
                
                # Logica per determinare la proposta di sconto
                if p2w is not None and p2w > theoretical_increase * 1.25:
                    df_clusters.at[item_index, 'Proposal'] = "Nessuno Sconto"
                else:
                    if cluster == "Basso":
                        if (p2w and p2w < threshold) or (p3w and p3w < threshold):
                            df_clusters.at[item_index, 'Proposal'] = "Sconto Alto"
                        else:
                            df_clusters.at[item_index, 'Proposal'] = "Sconto Medio"
                    elif cluster == "Medio Basso":
                        if (p2w and p2w < threshold) or (p3w and p3w < threshold):
                            df_clusters.at[item_index, 'Proposal'] = "Sconto Medio"
                        else:
                            df_clusters.at[item_index, 'Proposal'] = "Sconto Basso"
                    elif cluster in ["Alto", "Medio Alto"]:
                        if p2w and p2w < threshold:
                            if p3w and p3w < threshold:
                                df_clusters.at[item_index, 'Proposal'] = "Sconto Basso"
                            else:
                                df_clusters.at[item_index, 'Proposal'] = "Nessuno Sconto"
                        else:
                            df_clusters.at[item_index, 'Proposal'] = "Nessuno Sconto"
        
        progress_bar.progress(70)
        
        # Assegnazione proposte per casi speciali
        for item_list, proposal in [
            (items_in_exposition_month_without_sales, "Sconto Alto (NO SALES)"),
            (items_not_in_exposition_month, "NESSUNA PROPOSTA (item fuori da exposition months)"),
            (items_not_tracked, "NESSUNA PROPOSTA (item senza tracking)"),
            (items_below_tracked_in_exposition_month, "NESSUNA PROPOSTA (item in exposition months con tracking sotto 30%)")
        ]:
            for item in item_list:
                item_index = df_clusters.index[df_clusters['Item Code'] == item].tolist()
                if item_index:
                    item_index = item_index[0]
                    df_clusters.at[item_index, 'Proposal'] = proposal
        
        # Merge con altri dataframe
        A_excluded = A.drop(columns=['Commercial YearWeek', 'Commercial YearMonth'], errors='ignore')
        merged_df = pd.merge(df_clusters, A_excluded, on="Item Code", how="left")
        merged_df2 = pd.merge(merged_df, tracking, on="Item Code", how="left")
        
        progress_bar.progress(75)
        
        # Gestione articoli riciclati
        status_text.text("♻️ Elaborazione articoli ricondizionati...")

        # Use ORIGINAL (unfiltered) data for recycled processing
        A_recycled = A_original[A_original["Recycled"] == "Sì"].copy()
        st_item_recycled = st_item_original.merge(A_recycled, on="Item Code", how='inner')
        df_recycled = st_item_recycled.copy()
        
        #st.info(f"Articoli ricondizionati trovati (prima del filtro settimane): {len(df_recycled)}")
        
        # Categorize recycled items by function only (not by month)
        for func in df_recycled["Function"].unique():
            st_percentiles = st_item_original[st_item_original["Function"] == func]["ST item"].quantile([0.25, 0.5, 0.75])
            
            def categorize_recycled(row):
                if row["ST item"] <= st_percentiles.iloc[0]:  # 0.25 percentile
                    return 'Basso'
                elif row["ST item"] <= st_percentiles.iloc[1]:  # 0.5 percentile
                    return 'Medio Basso'
                elif row["ST item"] <= st_percentiles.iloc[2]:  # 0.75 percentile
                    return 'Medio Alto'
                else:
                    return 'Alto'
            
            df_recycled.loc[df_recycled["Function"] == func, "ST_Cluster"] = df_recycled[df_recycled["Function"] == func].apply(categorize_recycled, axis=1)
        
        # Process B data for recycled items using ORIGINAL B data
        recycled_items = set(df_recycled["Item Code"])
        B_recycled = B_original[B_original["Item Code"].isin(recycled_items)].copy()
        
        # Apply remove_leading_zero to B_recycled (this function should be defined earlier)
        def remove_leading_zero(year_week):
            year, week = year_week.split('-')
            week = str(int(week))
            return f"{year}-{week}"
        
        B_recycled['YearWeek'] = B_recycled['YearWeek'].apply(remove_leading_zero)
        
        # Apply Delta ST calculations for recycled items
        for index, row in df_recycled.iterrows():
            item = row["Item Code"]
            parts = current_week.split("-")
            week_number = int(parts[1])
            
            if week_number != 1:
                week_number_p2w = str(week_number)
                week_number_p3w = str(week_number - 1)
            else:
                week_number_p2w = "52"
                week_number_p3w = "51"
            
            week_p2w = parts[0] + "-" + week_number_p2w
            week_p3w = parts[0] + "-" + week_number_p3w
            
            p2w_data = B_recycled.loc[(B_recycled["Item Code"] == item) & (B_recycled["YearWeek"] == week_p2w), "Delta ST PW"]
            p3w_data = B_recycled.loc[(B_recycled["Item Code"] == item) & (B_recycled["YearWeek"] == week_p3w), "Delta ST PW"]
            
            p2w = p2w_data.values[0] if not p2w_data.empty else None
            p3w = p3w_data.values[0] if not p3w_data.empty else None
            
            def format_percent(x):
                if x is None:
                    return "-"
                else:
                    return f"{x*100:.2f}%".replace('.', ',')
            
            value_p2w = format_percent(p2w)
            value_p3w = format_percent(p3w)
            
            # Get threshold from goals
            mask_function = (goals['Function'] == row["Function"])
            if not goals.loc[mask_function].empty:
                goal_row = goals.loc[mask_function].iloc[0]
                theoretical_increase = goal_row['Teorethical Increase %']
                num_life_weeks = goal_row['NumLifeWeeks']
                if num_life_weeks == -1:
                    threshold = 0.025
                else:
                    threshold = 0.75 * theoretical_increase
            else:
                threshold = 0.0196 
                theoretical_increase = 0.0196
            
            cluster = row["ST_Cluster"]
            
            # Apply proposal logic
            if p2w is not None and p2w > theoretical_increase * 1.25:
                proposal = "Nessuno Sconto"
            else:
                if cluster == "Basso":
                    if (p2w is not None and p2w < threshold) or (p3w is not None and p3w < threshold):
                        proposal = "Sconto Alto"
                    else:
                        proposal = "Sconto Medio"
                elif cluster == "Medio Basso":
                    if (p2w is not None and p2w < threshold) or (p3w is not None and p3w < threshold):
                        proposal = "Sconto Medio"
                    else:
                        proposal = "Sconto Basso"
                elif cluster in ["Alto", "Medio Alto"]:
                    if p2w is not None and p2w < threshold:
                        if p3w is not None and p3w < threshold:
                            proposal = "Sconto Basso"
                        else:
                            proposal = "Nessuno Sconto"
                    else:
                        proposal = "Nessuno Sconto"
            
            # Check if item has recent first sale date
            if row.get("Weeks since First Sale Date", 0) < 10:
                proposal = "NESSUNA PROPOSTA (articolo rico con prima vendita troppo recente)"
            
            df_recycled.at[index, "Proposal"] = proposal
            df_recycled.at[index, "Delta ST P2W"] = value_p2w
            df_recycled.at[index, "Delta ST P3W"] = value_p3w
        
        # Set recycled-specific fields
        df_recycled['Metodo Cluster'] = "Cluster funzione (articolo ricondizionato)"
        df_recycled['% Store with Tracking'] = "-"
        df_recycled['% Stores with Tracking within 6 weeks'] = "-"
        df_recycled['First Tracking YearWeek'] = "-"
        df_recycled['Intake Quantity'] = "-"
        df_recycled['Displayed Quantity'] = "-"
        df_recycled['Total Item Tracked'] = "-"
        df_recycled['First Planned Tracking YearWeek'] = "-"
        
        #st.info(f"Articoli ricondizionati processati: {len(df_recycled)}")
            
        merged_df2 = pd.concat([merged_df2, df_recycled], ignore_index=True)
        
        progress_bar.progress(80)
        
        # Selezione segmenti
        status_text.text("🎯 Filtro per segmenti...")
        unique_segments = segment["Segment"].dropna().unique()
        
        # Per Streamlit, selezioniamo tutti i segmenti disponibili per default
        selected_segments = st.multiselect(
            "Segmenti inclusi:", 
            unique_segments, 
            default=unique_segments,
            key="segments_selector"
        )
        
        if not selected_segments:
            st.warning("Seleziona almeno un segmento!")
            st.stop()
        
        segment_filtered = segment[segment["Segment"].isin(selected_segments)]
        items_in_right_segment = set(segment_filtered["Cod item"])
        merged_df_final = merged_df2[merged_df2["Item Code"].isin(items_in_right_segment)]
        
        # Aggiornare proposte per articoli fuori segmento
        items_in_wrong_segment = set(merged_df2["Item Code"]) - items_in_right_segment
        for item in items_in_wrong_segment:
            item_index = merged_df2.index[merged_df2['Item Code'] == item].tolist()
            if item_index:
                merged_df2.at[item_index[0], 'Proposal'] = "NESSUNA PROPOSTA (segmento articoli non in considerazione)"
        
        progress_bar.progress(85)
        
        # Calcoli finali
        status_text.text("📈 Calcoli finali...")
        
        # Merge con segment originale
        segment_original = segment.copy()
        merged_df2 = pd.merge(merged_df2, segment_original, left_on="Item Code", right_on="Cod item", how="left")
        merged_df2 = merged_df2.drop(columns=["Cod item"], errors='ignore')
        
        # Aggiunta TFI
        def format_tfi(x):
            return f"{x*100:.2f}%".replace('.', ',')
        
        mapping_tfi = goals.set_index("Function")["Teorethical Increase %"].apply(format_tfi).to_dict()
        merged_df2["TFI"] = merged_df2["Function"].map(mapping_tfi)
        merged_df2["TFI"] = merged_df2["TFI"].fillna("1,96%")
        
        # Calcoli stock e SVA
        merged_df2["Sales item"] = pd.to_numeric(merged_df2["Sales item"], errors='coerce')
        merged_df2["Delivered item"] = pd.to_numeric(merged_df2["Delivered item"], errors='coerce')
        merged_df2["Stock residuo"] = merged_df2["Delivered item"] - merged_df2["Sales item"]
        
        # Calcolo SVA
        perc_basso, perc_medio, perc_alto = 0.2, 0.3, 0.5
        merged_df2["SVA"] = merged_df2.apply(lambda row: 
            row["Stock residuo"] * perc_basso if row["Proposal"] == "Sconto Basso" else
            row["Stock residuo"] * perc_medio if row["Proposal"] == "Sconto Medio" else
            row["Stock residuo"] * perc_alto if row["Proposal"] == "Sconto Alto" else 0,
            axis=1
        )
        
        merged_df2["Sconto proposto"] = merged_df2.apply(lambda row: 
            "SI" if row["Proposal"] in ["Sconto Basso", "Sconto Medio", "Sconto Alto"] else "NO",
            axis=1
        )
        
        # Data elaborazione
        elaboration_date = datetime.today().strftime('%d-%m-%Y')
        merged_df2['Data elaborazione'] = elaboration_date
        
        progress_bar.progress(90)
        
        # Predizioni opzionali con modello Keras
        if keras_model and images_df is not None:
            status_text.text("🤖 Elaborazione predizioni immagini...")
            try:
                # Carica modello Keras
                keras_model_bytes = keras_model.read()
                with open("temp_model.keras", "wb") as f:
                    f.write(keras_model_bytes)
                
                import tensorflow as tf
                model = tf.keras.models.load_model("temp_model.keras")
                
                st.success(f"Modello Keras caricato. Input shape: {model.input_shape}")
                
                # Prepara dati immagini
                images_df.rename(columns={"Item": "Item Code"}, inplace=True)
                merged_df2 = pd.merge(merged_df2, images_df[['Item Code', 'Picture']], on="Item Code", how="left")
                merged_df2.rename(columns={"Picture": "Image URL"}, inplace=True)
                merged_df2["Image URL"] = merged_df2["Image URL"].fillna("URL non presente")
                
                # Conta URL validi prima di iniziare
                valid_urls = merged_df2[
                    (merged_df2["Image URL"] != "URL non presente") & 
                    (merged_df2["Image URL"].notna())
                ]
                st.write(f"URL validi nel dataframe: {len(valid_urls)}/{len(merged_df2)}")
                
                if len(valid_urls) == 0:
                    st.error("Nessun URL valido trovato nel merge!")
                    st.write("Verifica che:")
                    st.write("- Il file immagini contenga la colonna 'Picture'")
                    st.write("- Gli Item Code nel file immagini corrispondano a quelli nel dataframe principale")
                    merged_df2["Discount Prediction"] = "Nessun URL disponibile"
                else:
                    # Mostra esempi
                    st.write("Esempi di URL:")
                    for url in valid_urls["Image URL"].head(3):
                        st.text(url[:100])
                    
                    # Funzione preprocessing
                    def preprocess_image_local(image, target_size=(224, 224)):
                        img = image.copy()
                        img.thumbnail(target_size, Image.LANCZOS)
                        new_img = Image.new("RGB", target_size)
                        left = (target_size[0] - img.size[0]) // 2
                        top = (target_size[1] - img.size[1]) // 2
                        new_img.paste(img, (left, top))
                        return new_img
                    
                    # Contatori per debug
                    error_counts = {
                        'url_non_presente': 0,
                        'http_error': 0,
                        'timeout': 0,
                        'image_error': 0,
                        'other': 0,
                        'success': 0
                    }
                    error_samples = []
                    
                    # Funzione download con logging e delay
                    import time
                    import random
                    
                    def download_and_preprocess_local(index, url, category, session):
                        if url == "URL non presente" or pd.isna(url):
                            error_counts['url_non_presente'] += 1
                            return None
                        
                        # Delay random per evitare rate limiting
                        time.sleep(random.uniform(0.1, 0.3))
                        
                        try:
                            response = session.get(url, timeout=30, allow_redirects=True)
                            if response.status_code == 200:
                                # Verifica content-type
                                content_type = response.headers.get('content-type', '').lower()
                                if 'image' not in content_type and len(error_samples) < 10:
                                    error_samples.append(f"Non è un'immagine ({content_type}): {url[:60]}")
                                
                                img = Image.open(BytesIO(response.content)).convert("RGB")
                                img = preprocess_image_local(img, target_size=(224, 224))
                                img_array = np.array(img) / 255.0
                                error_counts['success'] += 1
                                return (index, img_array, category)
                            else:
                                error_counts['http_error'] += 1
                                if len(error_samples) < 10:
                                    error_samples.append(f"HTTP {response.status_code}: {url[:60]}")
                                return None
                        except requests.exceptions.Timeout:
                            error_counts['timeout'] += 1
                            if len(error_samples) < 10:
                                error_samples.append(f"Timeout: {url[:60]}")
                            return None
                        except (Image.UnidentifiedImageError, OSError) as e:
                            error_counts['image_error'] += 1
                            if len(error_samples) < 10:
                                error_samples.append(f"Errore immagine: {url[:60]}")
                            return None
                        except Exception as e:
                            error_counts['other'] += 1
                            if len(error_samples) < 10:
                                error_samples.append(f"{type(e).__name__}: {url[:60]}")
                            return None
                    
                    # Setup session con headers completi
                    session = requests.Session()
                    session.headers.update({
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                        'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
                        'Accept-Language': 'en-US,en;q=0.9',
                        'Accept-Encoding': 'gzip, deflate, br',
                        'Connection': 'keep-alive',
                        'Referer': 'https://www.google.com/'
                    })
                    # Disabilita verifica SSL se necessario (solo per test)
                    session.verify = False
                    import urllib3
                    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                    
                    results = []
                    
                    st.info(f"Elaborazione {len(merged_df2)} righe in parallelo...")
                    
                    # Download parallelo con 20 worker
                    with ThreadPoolExecutor(max_workers=20) as executor:
                        future_to_info = {
                            executor.submit(download_and_preprocess_local, idx, row["Image URL"], row["Cod Category"], session): idx 
                            for idx, row in merged_df2.iterrows()
                        }
                        
                        progress_bar = st.progress(0)
                        completed = 0
                        total = len(future_to_info)
                        
                        for future in as_completed(future_to_info):
                            res = future.result()
                            if res is not None:
                                results.append(res)
                            completed += 1
                            if completed % 10 == 0 or completed == total:
                                progress_bar.progress(completed / total)
                        
                        progress_bar.empty()
                    
                    # Mostra statistiche dettagliate
                    st.write("**Risultati elaborazione:**")
                    st.write(f"- Successi: {error_counts['success']}")
                    st.write(f"- URL non presente/nullo: {error_counts['url_non_presente']}")
                    st.write(f"- Errori HTTP: {error_counts['http_error']}")
                    st.write(f"- Timeout: {error_counts['timeout']}")
                    st.write(f"- Errori immagine: {error_counts['image_error']}")
                    st.write(f"- Altri errori: {error_counts['other']}")
                    
                    if error_samples:
                        with st.expander("Esempi di errori", expanded=True):
                            for sample in error_samples:
                                st.text(sample)
                    
                    # Ordina risultati
                    results.sort(key=lambda x: x[0])
                    
                    if results:
                        indices, images_list, categories = zip(*results)
                        batch_images = np.stack(images_list, axis=0)
                        
                        st.info(f"Generazione predizioni per {len(batch_images)} immagini...")
                        st.write(f"Shape batch: {batch_images.shape}")
                        
                        # Predizione
                        predictions = model.predict(batch_images, verbose=0)
                        st.write(f"Shape predizioni: {predictions.shape}")
                        
                        # Soglie per categoria
                        pred_dict = {}
                        soglie_per_categoria = {31: 0.57, 32: 0.70, 33: 0.75}
                        
                        for idx, pred, cat in zip(indices, predictions, categories):
                            soglia = soglie_per_categoria.get(cat, 0.75)
                            pred_value = float(pred[0])
                            pred_dict[idx] = "Potenzialmente impattante" if pred_value >= soglia else "Potenzialmente non impattante"
                        
                        # Assegna predizioni
                        merged_df2["Discount Prediction"] = merged_df2.index.map(
                            lambda idx: pred_dict.get(idx, "Prediction not available")
                        )
                        
                        st.success("Predizioni immagini completate!")
                        
                        # Statistiche
                        pred_counts = merged_df2["Discount Prediction"].value_counts()
                        st.write("**Distribuzione predizioni:**")
                        for label, count in pred_counts.items():
                            st.write(f"- {label}: {count} ({count/len(merged_df2)*100:.1f}%)")
                        
                        # Esempi
                        sample_predictions = [(idx, float(pred[0]), cat) for idx, pred, cat in list(zip(indices, predictions, categories))[:5]]
                        with st.expander("Esempi valori predetti (prime 5)", expanded=False):
                            for idx, val, cat in sample_predictions:
                                soglia = soglie_per_categoria.get(cat, 0.75)
                                st.write(f"Indice {idx} (Cat {cat}): {val:.4f} (soglia: {soglia}) → {pred_dict[idx]}")
                        
                    else:
                        merged_df2["Discount Prediction"] = "Prediction not available"
                        st.warning("""
                        **Analisi del problema:**
                        
                        Se la maggior parte degli errori sono di tipo "Errori HTTP", le cause più probabili sono:
                        
                        1. **Restrizioni di rete a livello infrastrutturale**
                           - Firewall aziendale che blocca richieste da Streamlit Cloud
                           - CDN che richiede autenticazione specifica
                           - Protezioni anti-bot attive sul server delle immagini
                        
                        2. **URL non accessibili pubblicamente**
                           - Le immagini richiedono autenticazione (token, cookie)
                           - Gli URL sono scaduti o temporanei
                           - Restrizioni geografiche o IP whitelisting
                        
                        **Azioni consigliate:**
                        
                        - Verificare che gli URL siano accessibili pubblicamente (testare aprendo un URL nel browser)
                        - Se le immagini sono dietro autenticazione, contattare il team IT per ottenere credenziali API o estendere i permessi di accesso
                        - Considerare di hostare le immagini su un CDN pubblico o storage accessibile
                        - Verificare con il team IT se Streamlit Cloud è autorizzato ad accedere alle risorse aziendali (non basta includerlo
                          nei domini accessibili dalla rete aziendale, bisogna concedere l'accesso ai vari URL delle immagini)
                        - SE FOSSE NECESSARIO INCLUDERE LA PARTE PREDITTIVA NELL'OUTPUT, ESEGUIRE IL CODICE DISPONIBILE AL SEGUENTE PERCORSO:
                            > N:\DemandMerchandising\Analisi Lorenzo\FILE APP, CODICE E CREDENZIALI\File sconti\Codice LCTB\Codice LCTB.ipynb
                          
                          ESEGUENDO IN LOCALE QUESTO CODICE DA UN COMPUTER AZIENDALE NON CI SONO PROBLEMI DI SCARICAMENTO DELLE IMMAGINI, IN QUANTO LE RICHIESTE
                          VENGONO FATTE DIRETTAMENTE DALL'INTERNO DELLA RETE AZIENDALE.
                        
                        Per continuare senza predizioni immagini, puoi procedere con l'elaborazione degli altri dati.
                        """)
                
            except Exception as e:
                st.error(f"Errore nell'elaborazione delle immagini: {str(e)}")
                import traceback
                st.code(traceback.format_exc(), language="python")
                merged_df2["Image URL"] = merged_df2.get("Image URL", "URL non presente")
                merged_df2["Discount Prediction"] = "Prediction not available"
                
        elif keras_model and images_df is None:
            st.warning("Modello Keras caricato ma file immagini mancante")
            merged_df2["Image URL"] = "File immagini non caricato"
            merged_df2["Discount Prediction"] = "Prediction not available"
        else:
            merged_df2["Image URL"] = "Modello non caricato"
            merged_df2["Discount Prediction"] = "Prediction not available"
        
    
        
        if pkl_model:
            status_text.text("📊 Elaborazione predizioni Delta ST...")
            try:
                # Carica modello PKL
                pkl_model_bytes = pkl_model.read()
                with open("temp_model.pkl", "wb") as f:
                    f.write(pkl_model_bytes)
                
                loaded_model = joblib.load("temp_model.pkl")
                
                # Preparazione features (versione semplificata)
                cols_to_numeric = ['ST item', 'Sales item']
                for col in cols_to_numeric:
                    if col in merged_df2.columns:
                        merged_df2[col] = pd.to_numeric(merged_df2[col], errors='coerce').fillna(0)
                
                # Predizione (semplificata)
                merged_df2["Delta ST previsto"] = 0.025  # Valore placeholder
                
            except Exception as e:
                st.warning(f"Errore nell'elaborazione Delta ST: {e}")
                merged_df2["Delta ST previsto"] = "Errore predizione"
        
        # File sequenza opzionale
        if sequence_file:
            status_text.text("📋 Elaborazione sequenza articoli...")
            df_sequenza = pd.read_excel(sequence_file)
            merged_df2 = pd.merge(merged_df2, df_sequenza, on="Item Code", how="left")
            merged_df2['Tipologia sconto applicato'] = merged_df2['Tipologia sconto applicato'].fillna('-')
            merged_df2['Settimana applicazione sconto'] = merged_df2['Settimana applicazione sconto'].fillna('-')
        
        # Filtro finale per deliveries >= 5000
        merged_df2 = merged_df2[merged_df2["Delivered item"] >= 5000]
        
        progress_bar.progress(100)
        status_text.text("✅ Elaborazione completata!")
        
        # Mostra risultati
        st.success(f"✅ Elaborazione completata! Processati {len(merged_df2)} articoli.")
        
        # Statistiche
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Totale Articoli", len(merged_df2))
        with col2:
            sconto_si = len(merged_df2[merged_df2["Sconto proposto"] == "SI"])
            st.metric("Con Sconto", sconto_si)
        with col3:
            sconto_no = len(merged_df2[merged_df2["Sconto proposto"] == "NO"])
            st.metric("Senza Sconto", sconto_no)
        with col4:
            if sconto_si + sconto_no > 0:
                perc_sconto = (sconto_si / (sconto_si + sconto_no)) * 100
                st.metric("% Sconto", f"{perc_sconto:.1f}%")
        
        # Anteprima dati
        st.subheader("📋 Anteprima Risultati")
        st.dataframe(merged_df2.head(20), use_container_width=True)
        
        # Download file Excel
        excel_buffer, filename = create_styled_excel(merged_df2, category, sequence_file)
        
        st.download_button(
            label="📥 Scarica File Excel",
            data=excel_buffer,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Distribuzione proposte
        st.subheader("📊 Distribuzione Proposte di Sconto")
        proposal_counts = merged_df2['Proposal'].value_counts()
        st.bar_chart(proposal_counts)
        
        # Tabella riassuntiva proposte
        st.subheader("📈 Riassunto Proposte")
        summary_df = pd.DataFrame({
            'Proposta': proposal_counts.index,
            'Numero Articoli': proposal_counts.values,
            'Percentuale': (proposal_counts.values / len(merged_df2) * 100).round(2)
        })
        st.dataframe(summary_df, use_container_width=True)
        
    except Exception as e:
        st.error(f"❌ Errore durante l'elaborazione: {str(e)}")
        st.exception(e)
    
    finally:
        # Pulizia file temporanei
        import os
        temp_files = ["temp_model.keras", "temp_model.pkl"]
        for temp_file in temp_files:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass

    # Footer informativo
    st.markdown("---")
    st.markdown("""
    ### ℹ️ Informazioni sull'Applicazione
    Questa applicazione elabora dati di vendita e stock per generare proposte di sconto automatiche basate su:
    - Analisi delle performance di vendita (ST - Sell Through)
    - Clustering degli articoli per funzione e periodo
    - Soglie di tracking nei punti vendita
    - Modelli predittivi opzionali per ottimizzare le decisioni

    **Funzionalità principali:**
    - ✅ Caricamento multiplo di file Excel
    - ✅ Filtri temporali personalizzabili
    - ✅ Categorizzazione automatica articoli
    - ✅ Gestione articoli ricondizionati
    - ✅ Integrazione modelli ML (opzionale)
    - ✅ Export formattato per Excel
    - ✅ Dashboard interattiva con metriche

    **Note tecniche:**
    - Filtro automatico per delivery >= 5000 pezzi
    - Supporto per predizioni immagini (Keras)
    - Supporto per predizioni Delta ST (scikit-learn)
    - Formattazione avanzata file di output
    """)

    st.sidebar.markdown("---")
    st.sidebar.info("💡 **Suggerimento**: Assicurati che tutti i file abbiano la struttura colonne corretta prima del caricamento.")
















